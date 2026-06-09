import sys, json, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Color Palette ──────────────────────────────────────────────
C_TITLE_BG   = '0D2137'
C_TITLE_FG   = 'FAC775'
C_HEADER_BG  = '1B3A5C'
C_HEADER_FG  = 'FFFFFF'
C_SUB_BG     = '2E5F8E'
C_SUB_FG     = 'FFFFFF'
C_SUB2_BG    = '3A7ABE'
C_ACCENT1    = '000000'   # hitam – data angka (was hijau 5DCAA5)
C_ACCENT2    = 'F09595'   # merah – negatif
C_ACCENT3    = 'FAC775'   # gold
C_ROW_EVEN   = 'F0F6FF'
C_ROW_ODD    = 'FFFFFF'
C_TOTAL_BG   = 'D4E4F0'
C_BORDER     = 'B8C8D8'

IDR = '#,##0'

def rp(n):
    try: return int(n or 0)
    except: return 0

def thin_border():
    s = Side(style='thin', color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style='medium', color='8AABBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, val='', fmt=None, fg='000000', bg=None, bold=False,
          align='center', valign='center', wrap=False, border='thin', size=9):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name='Arial', size=size, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    cell.border = thin_border() if border == 'thin' else medium_border()
    if fmt: cell.number_format = fmt
    if bg:  cell.fill = PatternFill('solid', fgColor=bg)
    return cell

def header(ws, row, col, val, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True,
           size=9, wrap=True, align='center'):
    return _cell(ws, row, col, val, fg=fg, bg=bg, bold=bold, size=size,
                 wrap=wrap, align=align, border='medium')

def data(ws, row, col, val, fmt=None, fg='1B3A5C', bg=None,
         bold=False, align='center'):
    bg = bg or (C_ROW_EVEN if row % 2 == 0 else C_ROW_ODD)
    return _cell(ws, row, col, val, fmt=fmt, fg=fg, bg=bg, bold=bold,
                 align=align, border='thin')

def money_cell(ws, row, col, val, bg=None):
    bg = bg or (C_ROW_EVEN if row % 2 == 0 else C_ROW_ODD)
    fg = C_ACCENT1 if rp(val) >= 0 else C_ACCENT2
    return _cell(ws, row, col, rp(val), fmt=IDR, fg=fg, bg=bg,
                 align='right', border='thin')

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def set_col_width(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[get_column_letter(col) if isinstance(col, int) else col].width = w

# ── Collect all unique inventory item names across all loss logs ──
def get_all_loss_item_names(po_list):
    names = []
    seen = set()
    for po in po_list:
        for log in (po.get('lossLog') or []):
            for item in (log.get('items') or []):
                nm = item.get('nama', '')
                if nm and nm not in seen:
                    names.append(nm)
                    seen.add(nm)
    return names

# ── Collect all unique component names from bundleDef.komponen ──
def get_all_bundle_product_names(po_list, bundle_def=None):
    """
    Ambil nama KOMPONEN dari bundleDef (bukan nama bundle).
    Contoh: bundleDef = [{nama:'Regulator Bundle', komponen:[{nama:'Regulator',qty:1},{nama:'Airandom',qty:1}]}]
    → returns ['Regulator', 'Airandom']

    Fallback jika tidak ada komponen: gunakan nama bundle itu sendiri.
    Fallback kedua: ambil dari bundleDetail PO.
    """
    names = []
    seen = set()

    for bd in (bundle_def or []):
        komponen = bd.get('komponen') or []
        if komponen:
            # Bundle punya komponen → expand ke nama komponen
            for k in komponen:
                nm = k.get('nama', '') or k.get('itemNama', '')
                if nm and nm not in seen:
                    names.append(nm)
                    seen.add(nm)
        else:
            # Bundle tanpa komponen → nama bundle sendiri jadi kolom
            nm = bd.get('nama', '')
            if nm and nm not in seen:
                names.append(nm)
                seen.add(nm)

    # Fallback dari bundleDetail PO (kalau bundleDef kosong)
    if not names:
        for po in po_list:
            for bd in (po.get('bundleDetail') or []):
                nm = bd.get('produk', '') or bd.get('nama', '')
                if nm and nm not in seen:
                    names.append(nm)
                    seen.add(nm)

    return names

# ── Get all unique souvenir names ──
def get_all_souvenir_names(po_list):
    names = []
    seen = set()
    for po in po_list:
        for sv in (po.get('souvenir') or []):
            nm = sv.get('nama', '')
            if nm and nm not in seen:
                names.append(nm)
                seen.add(nm)
    return names

# ── Build loss summary per PO: {nama_barang: netLoss qty} ──
def po_loss_by_item(po):
    result = {}
    for log in (po.get('lossLog') or []):
        for item in (log.get('items') or []):
            nm = item.get('nama', '')
            nl = rp(item.get('netLoss', 0))
            if nm:
                result[nm] = result.get(nm, 0) + nl
    return result

# ── Build souvenir summary per PO: {nama: qty} ──
def po_souvenir_by_item(po):
    result = {}
    for sv in (po.get('souvenir') or []):
        nm = sv.get('nama', '')
        qty = rp(sv.get('qty', 0))
        if nm:
            result[nm] = result.get(nm, 0) + qty
    return result

# ══════════════════════════════════════════════════════════════════
# SHEET: Sales atau Nego (satu sheet per orang)
# Layout kolom:
#   A=No, B=Tgl, C=Konsumen, D=Alamat/Lokasi, E=Nego/Sales, F=Coll
#   G.. = per produk (bundleDef): Sold + Retur  (2 kolom per produk)
#   Penagihan: 1..max_cicilan
#   Perhitungan Komisi: Komisi rate, Value Komisi, Souvenir
#   Payout: Termin 1-N, Termin N+1-end
#   Minus Unit: per loss item + per souvenir
#   Total Minus
#   Total Termin: 1..max_cicilan
# ══════════════════════════════════════════════════════════════════
def build_sales_nego_sheet(wb, entitas, peran, po_list, settings, periode_label,
                            loss_item_names, souvenir_names, bundle_def=None):
    nama = entitas.get('nama', '?')
    safe_sheet = f"{peran[:1]}-{nama}"[:31]
    ws = wb.create_sheet(safe_sheet)
    ws.sheet_view.showGridLines = False

    komisi_rate = rp(entitas.get('komisiRate', 0))
    pct1 = settings.get('split_komisi_pct1', 60) / 100
    pct2 = settings.get('split_komisi_pct2', 40) / 100
    splitN1 = settings.get('split_termin1', 4)

    # Filter PO milik entitas ini
    field = 'sales' if peran == 'Sales' else 'nego'
    my_pos = [p for p in po_list if p.get(field) == nama]

    # Hitung max cicilan
    max_cicilan = max((len(p.get('cicilan', [])) for p in my_pos), default=7)
    max_cicilan = max(max_cicilan, 7)

    # ── Produk dinamis untuk kolom Sold/Retur: dari komponen bundleDef ──
    # Ini adalah nama komponen dalam bundle (misal Regulator, Airandom)
    prod_names = get_all_bundle_product_names(po_list, bundle_def)
    # loss_item_names dipakai khusus untuk kolom Minus Unit (barang yang diretur/hilang)

    # ── Layout kolom ──
    COL_NO   = 1
    COL_TGL  = 2
    COL_KONS = 3
    COL_AMAT = 4
    COL_REK  = 5
    COL_COLL = 6

    # Barang: 2 kolom per produk (Sold, Retur)
    col_prod_start = 7
    n_prod = len(prod_names)
    col_prod_end = col_prod_start + n_prod * 2 - 1 if n_prod else col_prod_start + 1

    # Penagihan langsung setelah kolom produk (tidak ada kolom Total)
    COL_CIC1     = col_prod_end + 1 if n_prod else col_prod_start + 2
    col_cic_last = COL_CIC1 + max_cicilan - 1

    # Perhitungan Komisi
    COL_K_KOMISI  = col_cic_last + 1
    COL_K_VALUE   = COL_K_KOMISI + 1
    COL_K_SOVENIR = COL_K_VALUE  + 1
    COL_K_END     = COL_K_SOVENIR

    # Payout
    COL_PAY1 = COL_K_END + 1
    COL_PAY2 = COL_PAY1  + 1

    # Minus Unit
    col_minus_start = COL_PAY2 + 1
    col_minus_end   = col_minus_start + len(loss_item_names) - 1 if loss_item_names else col_minus_start - 1
    col_sv_start    = col_minus_end + 1 if loss_item_names else col_minus_start
    col_sv_end      = col_sv_start + len(souvenir_names) - 1 if souvenir_names else col_sv_start - 1

    if souvenir_names:
        COL_TOTAL_MINUS = col_sv_end + 1
    elif loss_item_names:
        COL_TOTAL_MINUS = col_minus_end + 1
    else:
        COL_TOTAL_MINUS = col_minus_start

    COL_LAST = COL_TOTAL_MINUS
    NCOLS    = COL_LAST

    sl = get_column_letter

    # ── Row 1: Periode ──
    ws.row_dimensions[1].height = 18
    ws.cell(row=1, column=2, value='Periode').font = Font(name='Arial', size=9, color='555555')
    ws.cell(row=1, column=3, value=periode_label).font = Font(name='Arial', size=9, bold=True, color='1B3A5C')

    # ── Row 2: Nama ──
    ws.row_dimensions[2].height = 18
    ws.cell(row=2, column=2, value=f'Nama {peran}').font = Font(name='Arial', size=9, color='555555')
    ws.cell(row=2, column=3, value=nama).font = Font(name='Arial', size=9, bold=True, color='1B3A5C')

    # ── Row 3: Rate komisi ──
    ws.row_dimensions[3].height = 15
    ws.cell(row=3, column=2, value='Rate Komisi').font = Font(name='Arial', size=8, color='888888')
    c3 = ws.cell(row=3, column=3, value=komisi_rate)
    c3.font = Font(name='Arial', size=8, color='2E7D52')
    c3.number_format = IDR

    # ── Row 4-6: Header ──
    R_H1, R_H2, R_H3 = 4, 5, 6
    for rh in [R_H1, R_H2, R_H3]:
        ws.row_dimensions[rh].height = 22

    # Kolom tetap A-F (merge rows 4-6)
    for col, label in [
        (COL_NO,  'No'),
        (COL_TGL, 'Tanggal'),
        (COL_KONS,'Konsumen'),
        (COL_AMAT,'Alamat/Lokasi'),
        (COL_REK, 'Nego' if peran == 'Sales' else 'Sales'),
        (COL_COLL,'Collector'),
    ]:
        header(ws, R_H1, col, label, bg=C_HEADER_BG, size=9)
        merge(ws, R_H1, col, R_H3, col)

    # ── Header Barang Keluar: per produk (masing-masing 2 kolom: Sold + Retur) ──
    if n_prod > 0:
        # Group header "Barang Keluar" span semua kolom produk
        header(ws, R_H1, col_prod_start, 'Barang Keluar', bg=C_HEADER_BG, size=9)
        merge(ws, R_H1, col_prod_start, R_H1, col_prod_end)

        # Sub-header per produk (row 5, merge 2 kolom: Sold + Retur)
        for i, nm in enumerate(prod_names):
            c = col_prod_start + i * 2
            header(ws, R_H2, c, nm, bg=C_SUB_BG, size=8, wrap=True)
            merge(ws, R_H2, c, R_H2, c + 1)
            header(ws, R_H3, c,     'Sold',  bg=C_SUB2_BG, size=8)
            header(ws, R_H3, c + 1, 'Retur', bg=C_SUB2_BG, size=8)
    else:
        # Tidak ada produk: 2 kolom placeholder
        header(ws, R_H1, col_prod_start, 'Barang Keluar', bg=C_HEADER_BG, size=9)
        merge(ws, R_H1, col_prod_start, R_H2, col_prod_start + 1)
        header(ws, R_H3, col_prod_start,     'Sold',  bg=C_SUB_BG, size=8)
        header(ws, R_H3, col_prod_start + 1, 'Retur', bg=C_SUB_BG, size=8)

    # ── Penagihan (Rp per termin) ──
    header(ws, R_H1, COL_CIC1, 'Penagihan', bg=C_HEADER_BG, size=9)
    merge(ws, R_H1, COL_CIC1, R_H2, col_cic_last)
    for i in range(max_cicilan):
        header(ws, R_H3, COL_CIC1 + i, str(i + 1), bg=C_SUB_BG, size=8)

    # ── Perhitungan Komisi ──
    header(ws, R_H1, COL_K_KOMISI, 'Perhitungan Komisi', bg=C_HEADER_BG, size=9)
    merge(ws, R_H1, COL_K_KOMISI, R_H2, COL_K_END)
    header(ws, R_H3, COL_K_KOMISI,  'Komisi',       bg=C_SUB_BG, size=8)
    header(ws, R_H3, COL_K_VALUE,   'Value Komisi', bg=C_SUB_BG, size=8)
    header(ws, R_H3, COL_K_SOVENIR, 'Souvenir',     bg=C_SUB_BG, size=8)

    # ── Payout ──
    header(ws, R_H1, COL_PAY1, 'Payout', bg=C_HEADER_BG, size=9)
    merge(ws, R_H1, COL_PAY1, R_H2, COL_PAY2)
    header(ws, R_H3, COL_PAY1, f'Termin 1-{splitN1}',           bg=C_SUB2_BG, size=8)
    header(ws, R_H3, COL_PAY2, f'Termin {splitN1+1}-{max_cicilan}', bg=C_SUB2_BG, size=8)

    # ── Minus Unit ──
    if loss_item_names or souvenir_names:
        minus_span_end = col_sv_end if souvenir_names else col_minus_end
        header(ws, R_H1, col_minus_start, 'Minus Unit', bg=C_HEADER_BG, size=9)
        merge(ws, R_H1, col_minus_start, R_H2, minus_span_end)
        for i, nm in enumerate(loss_item_names):
            header(ws, R_H3, col_minus_start + i, nm, bg=C_ACCENT2, fg='1B1B1B', size=8)
        for i, nm in enumerate(souvenir_names):
            header(ws, R_H3, col_sv_start + i, f'Sv:{nm}', bg=C_ACCENT2, fg='1B1B1B', size=8)
    else:
        header(ws, R_H1, col_minus_start, 'Minus Unit', bg=C_HEADER_BG, size=9)
        merge(ws, R_H1, col_minus_start, R_H3, col_minus_start)

    # ── Total Minus ──
    header(ws, R_H1, COL_TOTAL_MINUS, 'Total\nMinus', bg=C_ACCENT2, fg='1B1B1B', size=9)
    merge(ws, R_H1, COL_TOTAL_MINUS, R_H3, COL_TOTAL_MINUS)

    # ── Helper: sold & retur per produk dari satu PO ──
    # Buat lookup bundleDef by nama untuk resolusi komponen
    _bdef_map = {bd.get('nama', ''): bd for bd in (bundle_def or [])}
    _prod_set  = set(prod_names)

    def po_prod_qty(po):
        """
        Return {nama_komponen: qty_total} dari bundleDetail PO.
        Expand bundleDetail → komponen sesuai bundleDef.komponen.
        Contoh: bundleDetail=[{produk:'Paket A', qty:99}]
                Paket A.komponen=[{nama:'Regulator',qty:1},{nama:'Airandom',qty:1}]
                → {Regulator:99, Airandom:99}
        """
        result = {}
        bd_list = po.get('bundleDetail') or []

        for bd in bd_list:
            bundle_nm  = bd.get('produk', '') or bd.get('nama', '')
            bundle_qty = rp(bd.get('qty', 0))
            bdef       = _bdef_map.get(bundle_nm)
            komponen   = (bdef.get('komponen') or []) if bdef else []

            if komponen:
                # Expand ke tiap komponen
                for k in komponen:
                    nm = k.get('nama', '') or k.get('itemNama', '')
                    qty_per = rp(k.get('qty', 1)) or 1
                    matched = nm if nm in _prod_set else next(
                        (p for p in prod_names if p.lower() == nm.lower()), None
                    )
                    if matched:
                        result[matched] = result.get(matched, 0) + bundle_qty * qty_per
            else:
                # Bundle tanpa komponen: nama bundle sendiri
                matched = bundle_nm if bundle_nm in _prod_set else next(
                    (p for p in prod_names
                     if p.lower() in bundle_nm.lower() or bundle_nm.lower() in p.lower()), None
                )
                if matched:
                    result[matched] = result.get(matched, 0) + bundle_qty

        # Fallback: bundleDetail kosong tapi ada bundle total
        if not result and prod_names and rp(po.get('bundle', 0)) > 0:
            result[prod_names[0]] = rp(po.get('bundle', 0))
        return result

    def po_retur_qty(po):
        """
        Return {nama_komponen: qty_kembali} dari lossLog PO.
        Retur = barang yang secara fisik kembali (kembaliTotal), bukan yang hilang (netLoss).
        """
        result = {}
        for log in (po.get('lossLog') or []):
            for item in (log.get('items') or []):
                nm = item.get('nama', '')
                # kembaliTotal = kembaliGood + kembaliReject (barang yang dikembalikan)
                qty = rp(item.get('kembaliTotal', 0)) or rp(item.get('kembaliGood', 0)) + rp(item.get('kembaliReject', 0))
                if qty <= 0:
                    continue
                matched = nm if nm in _prod_set else next(
                    (p for p in prod_names if p.lower() == nm.lower()), None
                )
                if matched:
                    result[matched] = result.get(matched, 0) + qty
        return result

    def get_coll(po):
        """Ambil nama collector: dari field coll PO atau dari cicilan yang punya collector."""
        c_po = po.get('coll', '')
        if c_po:
            return c_po
        for c in (po.get('cicilan') or []):
            if c.get('collector'):
                return c.get('collector')
        return ''

    # ── Data rows ──
    r = 7
    for idx, po in enumerate(my_pos):
        ws.row_dimensions[r].height = 16
        cicilan  = po.get('cicilan', [])
        bundle   = rp(po.get('bundle', 0))

        prod_sold  = po_prod_qty(po)
        prod_retur = po_retur_qty(po)
        loss_by_item = po_loss_by_item(po)
        sv_by_item   = po_souvenir_by_item(po)

        total_sold  = sum(prod_sold.values()) or bundle
        total_retur = sum(prod_retur.values())

        value_komisi  = total_sold * komisi_rate
        total_sv_cost = rp(po.get('totalSouvenir', 0))
        pay1 = value_komisi * pct1
        pay2 = value_komisi * pct2

        # Alamat: coba lokasi dulu (field di PO baru), fallback ke alamat/konsumen.alamat
        alamat = po.get('lokasi') or po.get('alamat') or ''

        # Collector
        coll_nm = get_coll(po)

        data(ws, r, COL_NO,   idx + 1)
        data(ws, r, COL_TGL,  po.get('tanggal', ''), align='left')
        data(ws, r, COL_KONS, po.get('konsumen', ''), align='left')
        data(ws, r, COL_AMAT, alamat, align='left')
        rekan = po.get('nego', '') if peran == 'Sales' else po.get('sales', '')
        data(ws, r, COL_REK,  rekan, align='left')
        data(ws, r, COL_COLL, coll_nm, align='left')

        # Kolom per produk
        for i, nm in enumerate(prod_names):
            c_sold  = col_prod_start + i * 2
            c_retur = c_sold + 1
            sq = prod_sold.get(nm, 0)
            rq = prod_retur.get(nm, 0)
            _cell(ws, r, c_sold,  sq if sq else '', fmt='0' if sq else None,
                  fg=C_ACCENT1, bg=(C_ROW_EVEN if r % 2 == 0 else C_ROW_ODD), align='center', size=9)
            _cell(ws, r, c_retur, rq if rq else '', fmt='0' if rq else None,
                  fg=C_ACCENT2 if rq else '888888', bg=(C_ROW_EVEN if r % 2 == 0 else C_ROW_ODD), align='center', size=9)

        # ── Penagihan: Rp per termin ──
        for i, c in enumerate(cicilan[:max_cicilan]):
            col_c = COL_CIC1 + i
            if c.get('status') == 'lunas':
                data(ws, r, col_c, rp(c.get('tagihan', 0)), fmt=IDR, fg=C_ACCENT1)
            elif c.get('status') == 'kurang':
                data(ws, r, col_c, rp(c.get('terbayar', 0)), fmt=IDR, fg='FAC775')
            else:
                data(ws, r, col_c, '')
        for i in range(len(cicilan), max_cicilan):
            data(ws, r, COL_CIC1 + i, '')

        # Perhitungan Komisi
        data(ws, r, COL_K_KOMISI,  komisi_rate,  fmt=IDR, fg='1B3A5C')
        data(ws, r, COL_K_VALUE,   value_komisi, fmt=IDR, fg=C_ACCENT1)
        data(ws, r, COL_K_SOVENIR, total_sv_cost if total_sv_cost else '',
             fmt=IDR if total_sv_cost else None)

        # Payout
        data(ws, r, COL_PAY1, int(pay1), fmt=IDR, fg=C_ACCENT1)
        data(ws, r, COL_PAY2, int(pay2), fmt=IDR, fg=C_ACCENT1)

        # Minus Unit
        total_minus = 0
        for i, nm in enumerate(loss_item_names):
            qty = loss_by_item.get(nm, 0)
            data(ws, r, col_minus_start + i, qty if qty else '',
                 fg=C_ACCENT2 if qty else '888888')
            total_minus += qty
        for i, nm in enumerate(souvenir_names):
            qty = sv_by_item.get(nm, 0)
            data(ws, r, col_sv_start + i, qty if qty else '',
                 fg=C_ACCENT2 if qty else '888888')
        data(ws, r, COL_TOTAL_MINUS, total_minus if total_minus else '',
             fg=C_ACCENT2 if total_minus else '888888')

        r += 1

    # ── Row TOTAL ──
    ws.row_dimensions[r].height = 18
    _cell(ws, r, COL_NO, 'TOTAL', fg='FFFFFF', bg=C_TOTAL_BG, bold=True, align='center', size=9)
    merge(ws, r, COL_NO, r, COL_COLL)

    data_start = 7
    data_end   = r - 1

    if data_end >= data_start:
        # Per produk total
        for i in range(n_prod):
            c_sold  = col_prod_start + i * 2
            c_retur = c_sold + 1
            _cell(ws, r, c_sold,  f'=SUM({sl(c_sold)}{data_start}:{sl(c_sold)}{data_end})',
                  fmt='0', bg=C_TOTAL_BG, bold=True, align='center', fg='1B3A5C')
            _cell(ws, r, c_retur, f'=SUM({sl(c_retur)}{data_start}:{sl(c_retur)}{data_end})',
                  fmt='0', bg=C_TOTAL_BG, bold=True, align='center', fg=C_ACCENT2)
        # Penagihan Rp totals
        for i in range(max_cicilan):
            col_c = COL_CIC1 + i
            _cell(ws, r, col_c, f'=SUM({sl(col_c)}{data_start}:{sl(col_c)}{data_end})',
                  fmt=IDR, bg=C_TOTAL_BG, bold=True, align='right', fg=C_ACCENT1)
        # Komisi
        _cell(ws, r, COL_K_VALUE,   f'=SUM({sl(COL_K_VALUE)}{data_start}:{sl(COL_K_VALUE)}{data_end})',
              fmt=IDR, bg=C_TOTAL_BG, bold=True, align='right', fg=C_ACCENT1)
        _cell(ws, r, COL_K_SOVENIR, f'=SUM({sl(COL_K_SOVENIR)}{data_start}:{sl(COL_K_SOVENIR)}{data_end})',
              fmt=IDR, bg=C_TOTAL_BG, bold=True, align='right', fg=C_ACCENT1)
        # Payout
        _cell(ws, r, COL_PAY1, f'=SUM({sl(COL_PAY1)}{data_start}:{sl(COL_PAY1)}{data_end})',
              fmt=IDR, bg=C_TOTAL_BG, bold=True, align='right', fg=C_ACCENT1)
        _cell(ws, r, COL_PAY2, f'=SUM({sl(COL_PAY2)}{data_start}:{sl(COL_PAY2)}{data_end})',
              fmt=IDR, bg=C_TOTAL_BG, bold=True, align='right', fg=C_ACCENT1)
        # Minus
        for i in range(len(loss_item_names)):
            col_m = col_minus_start + i
            _cell(ws, r, col_m, f'=SUM({sl(col_m)}{data_start}:{sl(col_m)}{data_end})',
                  fmt='0', bg=C_TOTAL_BG, bold=True, align='center', fg=C_ACCENT2)
        for i in range(len(souvenir_names)):
            col_sv = col_sv_start + i
            _cell(ws, r, col_sv, f'=SUM({sl(col_sv)}{data_start}:{sl(col_sv)}{data_end})',
                  fmt='0', bg=C_TOTAL_BG, bold=True, align='center', fg=C_ACCENT2)
        _cell(ws, r, COL_TOTAL_MINUS, f'=SUM({sl(COL_TOTAL_MINUS)}{data_start}:{sl(COL_TOTAL_MINUS)}{data_end})',
              fmt='0', bg=C_TOTAL_BG, bold=True, align='center', fg=C_ACCENT2)
    else:
        for col in range(col_prod_start, NCOLS + 1):
            _cell(ws, r, col, '', bg=C_TOTAL_BG)

    # ── Column widths ──
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20   # Alamat/Lokasi lebih lebar
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14   # Collector
    for i in range(n_prod):
        c = col_prod_start + i * 2
        ws.column_dimensions[sl(c)].width     = 8   # Sold
        ws.column_dimensions[sl(c + 1)].width = 8   # Retur
    for i in range(max_cicilan):
        ws.column_dimensions[sl(COL_CIC1 + i)].width = 13
    ws.column_dimensions[sl(COL_K_KOMISI)].width  = 12
    ws.column_dimensions[sl(COL_K_VALUE)].width   = 14
    ws.column_dimensions[sl(COL_K_SOVENIR)].width = 12
    ws.column_dimensions[sl(COL_PAY1)].width = 14
    ws.column_dimensions[sl(COL_PAY2)].width = 14
    for i in range(len(loss_item_names)):
        ws.column_dimensions[sl(col_minus_start + i)].width = 11
    for i in range(len(souvenir_names)):
        ws.column_dimensions[sl(col_sv_start + i)].width = 11
    ws.column_dimensions[sl(COL_TOTAL_MINUS)].width = 11

    ws.freeze_panes = ws.cell(row=7, column=COL_CIC1)
    return ws


# ══════════════════════════════════════════════════════════════════
# SHEET: Global Monitoring — 1 baris per sales per hari
# Layout: grouped by bulan, TGL di-merge per hari,
#         subtotal per hari, subtotal per bulan, grand total
# ══════════════════════════════════════════════════════════════════
def build_global_monitoring_sheet(wb, db, periode_label):
    ws = wb.create_sheet('Monitoring Global')
    ws.sheet_view.showGridLines = False

    trip_list  = db.get('tripList', [])
    po_list    = db.get('poList', [])
    settings   = db.get('settings', {})
    bundle_def = db.get('bundleDef', [])
    perusahaan = settings.get('perusahaan', {})
    nama_perus = perusahaan.get('nama', 'Inter Global Pratama')

    # ── Daftar produk dari bundleDef (nama bundle = kolom barang) ──
    # Fallback: dari bundleDetail di PO kalau bundleDef kosong
    barang_map = {}  # nama → harga
    for bd in bundle_def:
        nm = bd.get('nama', '')
        if nm and nm not in barang_map:
            barang_map[nm] = rp(bd.get('harga', 0))
    # Fallback dari bundleDetail PO
    if not barang_map:
        for po in po_list:
            for bd in (po.get('bundleDetail') or []):
                nm = bd.get('produk', '') or bd.get('nama', '')
                if nm and nm not in barang_map:
                    barang_map[nm] = rp(bd.get('harga', 0))
    all_barang = list(barang_map.keys())

    # ── Daftar souvenir dari PO ──
    souvenir_map = {}
    for po in po_list:
        for sv in (po.get('souvenir') or []):
            nm = sv.get('nama', '')
            if nm and nm not in souvenir_map:
                souvenir_map[nm] = rp(sv.get('harga', 0))
    all_souvenir = list(souvenir_map.keys())

    # ── Semua sales aktif ──
    all_sales_aktif = [e['nama'] for e in db.get('entitas', [])
                       if e.get('peran') == 'Sales' and e.get('aktifStatus')]

    # ── Index PO by (tanggal, sales) ──
    po_by_tgl_sales = {}
    for po in po_list:
        key = (po.get('tanggal', ''), po.get('sales', ''))
        po_by_tgl_sales.setdefault(key, []).append(po)

    # ── Index trip by tanggalRaw ──
    trip_by_raw = {}
    for trip in trip_list:
        raw = trip.get('tanggalRaw', '')
        if raw:
            trip_by_raw.setdefault(raw, []).append(trip)

    # ── Kumpulkan semua tanggal unik dari trip + fallback dari PO ──
    trip_dates_raw = set(t.get('tanggalRaw', '') for t in trip_list if t.get('tanggalRaw'))

    # Fallback: ambil tanggal dari PO jika tidak ada di trip
    # PO menyimpan tanggal dalam format "10 Mei 2026" (formatDateShort)
    # Konversi ke YYYY-MM-DD agar bisa di-sort dan di-match
    BULAN_RAW_MAP = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'Mei': 5, 'Jun': 6,
        'Jul': 7, 'Agu': 8, 'Sep': 9, 'Okt': 10, 'Nov': 11, 'Des': 12
    }
    def tgl_to_raw(tgl_str):
        """Konversi 'DD Mon YYYY' → 'YYYY-MM-DD'"""
        parts = str(tgl_str).strip().split()
        if len(parts) == 3:
            bln = BULAN_RAW_MAP.get(parts[1])
            if bln:
                try:
                    return f'{int(parts[2]):04d}-{bln:02d}-{int(parts[0]):02d}'
                except:
                    pass
        return None

    # Buat mapping tanggalDisplay → raw untuk PO
    po_tgl_display_to_raw = {}
    for po in po_list:
        tgl_disp = po.get('tanggal', '')
        raw = tgl_to_raw(tgl_disp)
        if raw and tgl_disp:
            po_tgl_display_to_raw[tgl_disp] = raw
            trip_dates_raw.add(raw)

    # Untuk lookup PO di dalam loop, bangun juga index by raw date
    po_by_raw_sales = {}
    for po in po_list:
        tgl_disp = po.get('tanggal', '')
        raw = po_tgl_display_to_raw.get(tgl_disp, '')
        if raw:
            key = (raw, po.get('sales', ''))
            po_by_raw_sales.setdefault(key, []).append(po)

    all_dates_raw = sorted(trip_dates_raw)

    def parse_raw(raw):
        try:
            parts = raw.split('-')
            return int(parts[0]), int(parts[1]), int(parts[2])
        except:
            return (0, 0, 0)

    from collections import defaultdict
    dates_by_month = defaultdict(list)
    for raw in all_dates_raw:
        y, m, d = parse_raw(raw)
        dates_by_month[(y, m)].append(raw)

    BULAN_ID = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']

    # ── Layout kolom ──
    COL_TGL   = 1
    COL_SALES = 2
    COL_NEGO  = 3
    COL_SUPIR = 4
    COL_SESI  = 5
    COL_KONS  = 6

    col_brg_start = 7
    col_brg_end   = col_brg_start + len(all_barang) * 2 - 1 if all_barang else col_brg_start - 1

    col_sv_start = col_brg_end + 1 if all_barang else col_brg_start
    col_sv_end   = col_sv_start + len(all_souvenir) * 2 - 1 if all_souvenir else col_sv_start - 1

    COL_TOT_BRG_QTY = (col_sv_end + 1) if all_souvenir else ((col_brg_end + 1) if all_barang else col_brg_start)
    COL_TOT_BRG_VAL = COL_TOT_BRG_QTY + 1
    COL_TOT_SV_QTY  = COL_TOT_BRG_VAL + 1
    COL_TOT_SV_VAL  = COL_TOT_SV_QTY + 1
    COL_KET         = COL_TOT_SV_VAL + 1
    NCOLS = COL_KET

    sl = get_column_letter

    def sum_formula(col, r_start, r_end):
        return f'=SUM({sl(col)}{r_start}:{sl(col)}{r_end})'

    def write_subtotal_row(r, label, r_start, r_end, bg, label_col_end=COL_KONS):
        ws.row_dimensions[r].height = 16
        _cell(ws, r, COL_TGL, label, fg='FFFFFF', bg=bg, bold=True, align='left', size=9)
        merge(ws, r, COL_TGL, r, label_col_end)
        for i in range(len(all_barang)):
            c = col_brg_start + i * 2
            _cell(ws, r, c,   sum_formula(c,   r_start, r_end), fmt='0',  bg=bg, bold=True, align='center', fg='FFFFFF')
            _cell(ws, r, c+1, sum_formula(c+1, r_start, r_end), fmt=IDR,  bg=bg, bold=True, align='right',  fg='FAC775')
        for i in range(len(all_souvenir)):
            c = col_sv_start + i * 2
            _cell(ws, r, c,   sum_formula(c,   r_start, r_end), fmt='0',  bg=bg, bold=True, align='center', fg='FFFFFF')
            _cell(ws, r, c+1, sum_formula(c+1, r_start, r_end), fmt=IDR,  bg=bg, bold=True, align='right',  fg='FAC775')
        _cell(ws, r, COL_TOT_BRG_QTY, sum_formula(COL_TOT_BRG_QTY, r_start, r_end), fmt='0',  bg=bg, bold=True, align='center', fg='FFFFFF')
        _cell(ws, r, COL_TOT_BRG_VAL, sum_formula(COL_TOT_BRG_VAL, r_start, r_end), fmt=IDR,  bg=bg, bold=True, align='right',  fg='FAC775')
        _cell(ws, r, COL_TOT_SV_QTY,  sum_formula(COL_TOT_SV_QTY,  r_start, r_end), fmt='0',  bg=bg, bold=True, align='center', fg='FFFFFF')
        _cell(ws, r, COL_TOT_SV_VAL,  sum_formula(COL_TOT_SV_VAL,  r_start, r_end), fmt=IDR,  bg=bg, bold=True, align='right',  fg='FAC775')
        _cell(ws, r, COL_KET, '', bg=bg)

    # ── Row 1-2: Title ──
    ws.row_dimensions[1].height = 30
    cell = ws.cell(row=1, column=1, value=f'LAPORAN KEUANGAN GLOBAL — {nama_perus.upper()}')
    cell.font = Font(name='Arial', bold=True, size=14, color=C_TITLE_FG)
    cell.fill = PatternFill('solid', fgColor=C_TITLE_BG)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    merge(ws, 1, 1, 1, NCOLS)

    ws.row_dimensions[2].height = 16
    cell2 = ws.cell(row=2, column=1, value=f'Periode: {periode_label}')
    cell2.font = Font(name='Arial', size=9, color='888888')
    merge(ws, 2, 1, 2, NCOLS)

    # ── Row 3-5: Header ──
    R_H1, R_H2, R_H3 = 3, 4, 5
    ws.row_dimensions[R_H1].height = 24
    ws.row_dimensions[R_H2].height = 16
    ws.row_dimensions[R_H3].height = 18

    for col, label in [(COL_TGL,'Tanggal'), (COL_SALES,'Sales'), (COL_NEGO,'Nego'),
                        (COL_SUPIR,'Supir'), (COL_SESI,'Sesi'), (COL_KONS,'Konsumen/Tujuan')]:
        header(ws, R_H1, col, label, bg=C_HEADER_BG, size=9)
        merge(ws, R_H1, col, R_H3, col)

    if all_barang:
        header(ws, R_H1, col_brg_start, 'Barang Keluar', bg=C_HEADER_BG, size=9)
        merge(ws, R_H1, col_brg_start, R_H1, col_brg_end)
        for i, nm in enumerate(all_barang):
            c = col_brg_start + i * 2
            hrg = barang_map.get(nm, 0)
            lbl = f'{nm}\n@ {hrg:,}' if hrg else nm
            header(ws, R_H2, c, lbl, bg=C_SUB_BG, size=8, wrap=True)
            merge(ws, R_H2, c, R_H2, c+1)
            header(ws, R_H3, c,   'Unit', bg=C_SUB2_BG, size=8)
            header(ws, R_H3, c+1, 'Value (Rp)', bg=C_SUB2_BG, size=8)

    if all_souvenir:
        header(ws, R_H1, col_sv_start, 'Souvenir', bg='2E5E3E', size=9)
        merge(ws, R_H1, col_sv_start, R_H1, col_sv_end)
        for i, nm in enumerate(all_souvenir):
            c = col_sv_start + i * 2
            hrg = souvenir_map.get(nm, 0)
            lbl = f'{nm}\n@ {hrg:,}' if hrg else nm
            header(ws, R_H2, c, lbl, bg='3A7D52', size=8, wrap=True)
            merge(ws, R_H2, c, R_H2, c+1)
            header(ws, R_H3, c,   'Qty',        bg='4A9D62', size=8)
            header(ws, R_H3, c+1, 'Value (Rp)', bg='4A9D62', size=8)

    for col, label in [(COL_TOT_BRG_QTY,'Total\nUnit'), (COL_TOT_BRG_VAL,'Total\nValue'),
                        (COL_TOT_SV_QTY,'Sv\nQty'), (COL_TOT_SV_VAL,'Sv\nValue'),
                        (COL_KET,'Keterangan')]:
        header(ws, R_H1, col, label, bg=C_HEADER_BG, size=8)
        merge(ws, R_H1, col, R_H3, col)

    # ── Data rows ──
    r = 6
    grand_data_rows = []

    sorted_months = sorted(dates_by_month.keys())

    for (year, month) in sorted_months:
        month_dates = sorted(dates_by_month[(year, month)])
        nama_bulan  = BULAN_ID[month] if month <= 12 else str(month)

        ws.row_dimensions[r].height = 20
        cell_bln = ws.cell(row=r, column=1, value=f'  {nama_bulan.upper()} {year}')
        cell_bln.font = Font(name='Arial', bold=True, size=10, color=C_TITLE_FG)
        cell_bln.fill = PatternFill('solid', fgColor=C_TITLE_BG)
        cell_bln.alignment = Alignment(horizontal='left', vertical='center')
        merge(ws, r, 1, r, NCOLS)
        r += 1

        month_data_start = r

        for raw_date in month_dates:
            trips_hari = trip_by_raw.get(raw_date, [])
            tgl_display = next(
                (t.get('tanggal', '') for t in trips_hari if t.get('tanggal')),
                raw_date
            )
            # Jika tgl_display masih format raw (YYYY-MM-DD), cari dari PO
            if tgl_display == raw_date:
                for po in po_list:
                    po_raw = po_tgl_display_to_raw.get(po.get('tanggal', ''), '')
                    if po_raw == raw_date and po.get('tanggal'):
                        tgl_display = po.get('tanggal')
                        break

            # Sales yang hadir: gabung dari salesNama trip (bisa list atau string)
            sales_hadir = set()
            for trip in trips_hari:
                snama = trip.get('salesNama', trip.get('sales', ''))
                if isinstance(snama, list):
                    for s in snama: sales_hadir.add(s)
                elif snama:
                    sales_hadir.add(snama)
            # Fallback: jika tidak ada trip, ambil sales dari PO hari ini
            if not trips_hari:
                for po in po_list:
                    po_raw = po_tgl_display_to_raw.get(po.get('tanggal', ''), '')
                    if po_raw == raw_date and po.get('sales'):
                        sales_hadir.add(po.get('sales'))

            # Semua sales yang tampil hari ini
            sales_hari = sorted(sales_hadir) + [s for s in all_sales_aktif if s not in sales_hadir]

            day_row_start = r

            for sales_nm in sales_hari:
                ws.row_dimensions[r].height = 15
                is_off = sales_nm not in sales_hadir

                # Cari trip untuk sales ini
                trip_sales = None
                for t in trips_hari:
                    snama = t.get('salesNama', t.get('sales', ''))
                    names = snama if isinstance(snama, list) else ([snama] if snama else [])
                    if sales_nm in names:
                        trip_sales = t
                        break

                # Supir dari trip
                _supir_raw = trip_sales.get('supirNama', trip_sales.get('supir', '')) if trip_sales else ''
                supir = ', '.join(_supir_raw) if isinstance(_supir_raw, list) else str(_supir_raw or '')

                # PO milik sales ini hari ini
                # Coba lookup by (tgl_display, sales_nm) dulu
                my_pos = po_by_tgl_sales.get((tgl_display, sales_nm), [])
                # Fallback: lookup by (raw_date, sales_nm) jika tgl_display tidak match
                if not my_pos:
                    my_pos = po_by_raw_sales.get((raw_date, sales_nm), [])

                # Nego, Sesi, Konsumen: ambil dari PO (bukan trip)
                if my_pos:
                    # Bisa >1 PO, ambil nego unik
                    negos  = list(dict.fromkeys(po.get('nego', '')  for po in my_pos if po.get('nego')))
                    sesis  = list(dict.fromkeys(str(po.get('sesi', '')) for po in my_pos if po.get('sesi')))
                    konss  = list(dict.fromkeys(po.get('konsumen', '') for po in my_pos if po.get('konsumen')))
                    nego   = ', '.join(negos)
                    sesi   = ', '.join(sesis)
                    kons   = ', '.join(konss)
                elif trip_sales:
                    # Tidak ada PO tapi ada trip: pakai tujuan dari trip
                    nego  = ''
                    sesi  = ''
                    kons  = trip_sales.get('tujuan', '')
                else:
                    nego = sesi = kons = ''

                ket = 'OFF' if is_off else ''

                bg = 'F5F5F5' if is_off else (C_ROW_EVEN if r % 2 == 0 else C_ROW_ODD)
                fg_txt = 'AAAAAA' if is_off else '1B3A5C'

                _cell(ws, r, COL_TGL,   tgl_display, fg=fg_txt, bg=bg, align='center', size=9)
                _cell(ws, r, COL_SALES, sales_nm, fg='AA3333' if is_off else fg_txt, bg=bg, bold=is_off, align='left', size=9)
                _cell(ws, r, COL_NEGO,  nego,  fg=fg_txt, bg=bg, align='left',   size=9)
                _cell(ws, r, COL_SUPIR, supir, fg=fg_txt, bg=bg, align='left',   size=9)
                _cell(ws, r, COL_SESI,  sesi,  fg=fg_txt, bg=bg, align='center', size=9)
                _cell(ws, r, COL_KONS,  kons,  fg=fg_txt, bg=bg, align='left',   size=9)

                total_brg_qty = 0
                total_brg_val = 0
                if not is_off:
                    brg_q = {nm: 0 for nm in all_barang}
                    brg_v = {nm: 0 for nm in all_barang}
                    sv_q  = {nm: 0 for nm in all_souvenir}
                    sv_v  = {nm: 0 for nm in all_souvenir}

                    for po in my_pos:
                        bundle_total = rp(po.get('bundle', 0))
                        bd_list = po.get('bundleDetail') or []
                        matched_any = False

                        for bd in bd_list:
                            prod_nm  = bd.get('produk', '') or bd.get('nama', '')
                            qty_bd   = rp(bd.get('qty', 0))
                            harga_bd = rp(bd.get('harga', 0))
                            # Exact match dulu, lalu partial
                            matched = None
                            if prod_nm in barang_map:
                                matched = prod_nm
                            else:
                                matched = next(
                                    (nm for nm in all_barang
                                     if nm.lower() in prod_nm.lower() or prod_nm.lower() in nm.lower()),
                                    None
                                )
                            if matched:
                                brg_q[matched] += qty_bd
                                brg_v[matched] += qty_bd * (harga_bd or barang_map.get(matched, 0))
                                matched_any = True

                        # Fallback: bundle tidak punya bundleDetail atau tidak match
                        if not matched_any and bundle_total > 0 and all_barang:
                            nm0 = all_barang[0]
                            brg_q[nm0] += bundle_total
                            brg_v[nm0] += bundle_total * barang_map.get(nm0, 0)

                        for sv in (po.get('souvenir') or []):
                            nm  = sv.get('nama', '')
                            qty = rp(sv.get('qty', 0))
                            hrg = rp(sv.get('harga', 0)) or souvenir_map.get(nm, 0)
                            if nm in sv_q:
                                sv_q[nm] += qty
                                sv_v[nm] += qty * hrg

                    for i, nm in enumerate(all_barang):
                        c   = col_brg_start + i * 2
                        qty = brg_q.get(nm, 0)
                        val = brg_v.get(nm, 0)
                        _cell(ws, r, c,   qty if qty else '', fmt='0' if qty else None,
                              fg='000000' if qty else '888888', bg=bg, align='center', size=9)
                        _cell(ws, r, c+1, val if val else '', fmt=IDR if val else None,
                              fg='000000' if val else '888888', bg=bg, align='right', size=9)
                        total_brg_qty += qty
                        total_brg_val += val

                    total_sv_qty = 0
                    total_sv_val = 0
                    for i, nm in enumerate(all_souvenir):
                        c   = col_sv_start + i * 2
                        qty = sv_q.get(nm, 0)
                        val = sv_v.get(nm, 0)
                        _cell(ws, r, c,   qty if qty else '', fmt='0' if qty else None,
                              fg='000000' if qty else '888888', bg=bg, align='center', size=9)
                        _cell(ws, r, c+1, val if val else '', fmt=IDR if val else None,
                              fg='000000' if val else '888888', bg=bg, align='right', size=9)
                        total_sv_qty += qty
                        total_sv_val += val

                    _cell(ws, r, COL_TOT_BRG_QTY, total_brg_qty or '', fmt='0' if total_brg_qty else None,
                          fg='000000', bg=bg, bold=True, align='center', size=9)
                    _cell(ws, r, COL_TOT_BRG_VAL, total_brg_val or '', fmt=IDR if total_brg_val else None,
                          fg='000000', bg=bg, bold=True, align='right', size=9)
                    _cell(ws, r, COL_TOT_SV_QTY, total_sv_qty or '', fmt='0' if total_sv_qty else None,
                          fg='000000', bg=bg, align='center', size=9)
                    _cell(ws, r, COL_TOT_SV_VAL, total_sv_val or '', fmt=IDR if total_sv_val else None,
                          fg='000000', bg=bg, align='right', size=9)
                else:
                    for col in range(col_brg_start, NCOLS):
                        _cell(ws, r, col, '', bg=bg)

                _cell(ws, r, COL_KET, ket, fg='AA3333' if is_off else fg_txt, bg=bg, align='left', size=9)
                grand_data_rows.append(r)
                r += 1

            # Merge kolom TGL untuk semua sales hari ini
            n_rows = r - day_row_start
            if n_rows > 1:
                merge(ws, day_row_start, COL_TGL, r - 1, COL_TGL)

            # Subtotal per hari
            if n_rows > 0:
                write_subtotal_row(r, f'  Sub-Total {tgl_display}', day_row_start, r - 1, C_SUB_BG)
                r += 1

        # Subtotal per bulan
        write_subtotal_row(r, f'  TOTAL {nama_bulan.upper()} {year}', month_data_start, r - 1, C_HEADER_BG)
        r += 2

    # Grand Total
    if grand_data_rows:
        ws.row_dimensions[r].height = 20
        _cell(ws, r, COL_TGL, 'GRAND TOTAL', fg=C_TITLE_FG, bg=C_TITLE_BG, bold=True, align='left', size=10)
        merge(ws, r, COL_TGL, r, COL_KONS)
        for i in range(len(all_barang)):
            c = col_brg_start + i * 2
            _cell(ws, r, c,   sum_formula(c,   6, r-1), fmt='0',  bg=C_TITLE_BG, bold=True, align='center', fg='FAC775')
            _cell(ws, r, c+1, sum_formula(c+1, 6, r-1), fmt=IDR,  bg=C_TITLE_BG, bold=True, align='right',  fg='FAC775')
        for i in range(len(all_souvenir)):
            c = col_sv_start + i * 2
            _cell(ws, r, c,   sum_formula(c,   6, r-1), fmt='0',  bg=C_TITLE_BG, bold=True, align='center', fg='FAC775')
            _cell(ws, r, c+1, sum_formula(c+1, 6, r-1), fmt=IDR,  bg=C_TITLE_BG, bold=True, align='right',  fg='FAC775')
        _cell(ws, r, COL_TOT_BRG_QTY, sum_formula(COL_TOT_BRG_QTY, 6, r-1), fmt='0',  bg=C_TITLE_BG, bold=True, align='center', fg='FAC775')
        _cell(ws, r, COL_TOT_BRG_VAL, sum_formula(COL_TOT_BRG_VAL, 6, r-1), fmt=IDR,  bg=C_TITLE_BG, bold=True, align='right',  fg='FAC775')
        _cell(ws, r, COL_TOT_SV_QTY,  sum_formula(COL_TOT_SV_QTY,  6, r-1), fmt='0',  bg=C_TITLE_BG, bold=True, align='center', fg='FAC775')
        _cell(ws, r, COL_TOT_SV_VAL,  sum_formula(COL_TOT_SV_VAL,  6, r-1), fmt=IDR,  bg=C_TITLE_BG, bold=True, align='right',  fg='FAC775')
        _cell(ws, r, COL_KET, '', bg=C_TITLE_BG)

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 20
    for i in range(len(all_barang)):
        c = col_brg_start + i * 2
        ws.column_dimensions[get_column_letter(c)].width = 8
        ws.column_dimensions[get_column_letter(c+1)].width = 14
    for i in range(len(all_souvenir)):
        c = col_sv_start + i * 2
        ws.column_dimensions[get_column_letter(c)].width = 8
        ws.column_dimensions[get_column_letter(c+1)].width = 13
    for col in [COL_TOT_BRG_QTY, COL_TOT_BRG_VAL, COL_TOT_SV_QTY, COL_TOT_SV_VAL]:
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions[get_column_letter(COL_KET)].width = 14

    ws.freeze_panes = f'{get_column_letter(COL_SALES + 1)}6'
    return ws




def build_excel(db, periode_label, only_nama=None):
    wb = Workbook()
    wb.remove(wb.active)

    settings  = db.get('settings', {})
    po_list   = db.get('poList', [])
    entitas   = db.get('entitas', [])
    bundle_def = db.get('bundleDef', [])

    # Kumpulkan nama barang dari lossLog semua PO
    loss_item_names   = get_all_loss_item_names(po_list)
    souvenir_names    = get_all_souvenir_names(po_list)

    # Nama produk yang benar-benar dijual (dari bundleDef + bundleDetail PO)
    # Dipakai untuk kolom dinamis di sheet Sales/Nego (Sold per produk)
    bundle_product_names = get_all_bundle_product_names(po_list, bundle_def)

    # Jika lossLog belum ada (PO baru), pakai bundleDetail/bundleDef sebagai kolom barang
    effective_item_names = loss_item_names if loss_item_names else bundle_product_names

    # ── Sheet per Sales ──
    sales_list = [e for e in entitas if e.get('peran') == 'Sales' and (not only_nama or e.get('nama') == only_nama)]
    for e in sales_list:
        build_sales_nego_sheet(wb, e, 'Sales', po_list, settings,
                               periode_label, effective_item_names, souvenir_names,
                               bundle_def=bundle_def)

    # ── Sheet per Nego ──
    nego_list = [e for e in entitas if e.get('peran') == 'Nego' and (not only_nama or e.get('nama') == only_nama)]
    for e in nego_list:
        build_sales_nego_sheet(wb, e, 'Nego', po_list, settings,
                               periode_label, effective_item_names, souvenir_names,
                               bundle_def=bundle_def)

    # ── Sheet Global Monitoring ──
    if not only_nama:
        build_global_monitoring_sheet(wb, db, periode_label)

    # Freeze & set active
    wb.active = wb.worksheets[0]

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


if __name__ == '__main__':
    payload = json.load(sys.stdin)
    db      = payload.get('db', payload)
    periode = payload.get('periode', 'Semua Periode')

    filter_nama     = None
    monitoring_only = False

    if periode.startswith('__filter__'):
        rest        = periode[len('__filter__'):]
        parts       = rest.split('__', 1)
        filter_nama = parts[0]
        periode     = parts[1] if len(parts) > 1 else 'Semua Periode'
    elif periode.startswith('__monitoring__'):
        monitoring_only = True
        periode         = periode[len('__monitoring__'):]

    print(f'[Python] filter={filter_nama!r} monitoring={monitoring_only} periode={periode!r}', file=sys.stderr)

    if monitoring_only:
        from openpyxl import Workbook as _WB
        import io as _io
        _wb = _WB()
        _wb.remove(_wb.active)
        build_global_monitoring_sheet(_wb, db, periode)
        _wb.active = _wb.worksheets[0]
        _out = _io.BytesIO()
        _wb.save(_out)
        _out.seek(0)
        result = _out.getvalue()
    else:
        result = build_excel(db, periode, only_nama=filter_nama)

    sys.stdout.buffer.write(result)